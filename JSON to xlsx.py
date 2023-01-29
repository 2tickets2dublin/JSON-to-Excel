import json
import openpyxl as op
from openpyxl.styles import PatternFill

# converting json to dictionary
with open("db.json", "r") as file:
    dic = json.load(file)

# creating a list with movie data
movie_lst = []

for item in dic['movies']:
    movie_lst.append([item['year'], item['title'], item['genres']])

movie_lst = sorted(movie_lst)

# creating an Excel file and filling it with data
Fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

mov_file = op.Workbook()
mov_file_lst = mov_file.active

mov_file_lst.title = "Movies Database"
mov_file_lst["A1"] = "Film release year"
mov_file_lst["A1"].fill = Fill
mov_file_lst["B1"] = "Movie title"
mov_file_lst["B1"].fill = Fill
mov_file_lst["C1"] = "Movie genre"
mov_file_lst["C1"].fill = Fill

row = 2
for item in movie_lst:
    mov_file_lst[row][0].value = item[0]
    mov_file_lst[row][1].value = item[1]
    mov_file_lst[row][2].value = f"{item[2][0]}, {item[2][0]}"
    row += 1

#saving an Excel file
mov_file.save("mov.xlsx")
mov_file.close()